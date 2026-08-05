"""存档导出（Excel / PDF）+ 首页未写提醒。"""
from io import BytesIO

from flask import Blueprint, send_file, jsonify
from flask_login import login_required, current_user

from extensions import db
from models.class_student import Klass, Student, Enrollment
from models.lesson import Lesson, Review

reports_bp = Blueprint("reports", __name__, url_prefix="/reports")

# 课评状态 → 中文
STATUS_CN = {
    "pending": "待生成",
    "generating": "生成中",
    "draft": "草稿",
    "confirmed": "已确认",
    "failed": "生成失败",
    "leave": "请假",
}


def _class_scope(class_id):
    """取班级 + 在册学生 + 课次 + 课评映射（严格租户隔离）。"""
    klass = Klass.query.filter_by(
        id=class_id, user_id=current_user.id, deleted_at=None
    ).first_or_404()

    students = (
        Student.query.join(Enrollment, Enrollment.student_id == Student.id)
        .filter(
            Enrollment.class_id == class_id,
            Enrollment.deleted_at.is_(None),
            Student.deleted_at.is_(None),
        )
        .order_by(Student.name)
        .all()
    )

    lessons = (
        Lesson.query.filter_by(class_id=class_id, user_id=current_user.id)
        .filter(Lesson.deleted_at.is_(None))
        .order_by(Lesson.lesson_date.asc(), Lesson.created_at.asc())
        .all()
    )

    reviews = (
        Review.query.filter_by(class_id=class_id, user_id=current_user.id)
        .filter(Review.deleted_at.is_(None))
        .all()
    )
    # (lesson_id, student_id) -> Review
    rev_map = {(r.lesson_id, r.student_id): r for r in reviews}

    return klass, students, lessons, rev_map


def _lesson_label(lesson, idx):
    """课次显示名：优先标题，其次日期，最后序号。"""
    if lesson is None:
        return "（未关联课次）"
    title = (lesson.title or "").strip()
    date_s = lesson.lesson_date.strftime("%Y-%m-%d") if lesson.lesson_date else ""
    if title and date_s:
        return f"第{idx}次 {title}（{date_s}）"
    if title:
        return f"第{idx}次 {title}"
    if date_s:
        return f"第{idx}次 {date_s}"
    return f"第{idx}次"


@reports_bp.route("/export-xlsx/<class_id>", methods=["GET"])
@login_required
def export_xlsx(class_id):
    """导出 Excel：学生 × 课次 全矩阵，含课评状态与正文。"""
    klass, students, lessons, rev_map = _class_scope(class_id)

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # —— Sheet1：明细（一行 = 一个学生的一节课）——
    ws = wb.active
    ws.title = "课评明细"
    header = ["学生", "课次", "上课日期", "课评状态", "课评内容", "字数"]
    ws.append(header)

    head_fill = PatternFill("solid", fgColor="4F46E5")
    for col in range(1, len(header) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    if lessons:
        for idx, lesson in enumerate(lessons, start=1):
            label = _lesson_label(lesson, idx)
            date_s = (
                lesson.lesson_date.strftime("%Y-%m-%d") if lesson.lesson_date else ""
            )
            for s in students:
                r = rev_map.get((lesson.id, s.id))
                content = (r.content or "") if r else ""
                status = STATUS_CN.get(r.status, r.status) if r else "未生成"
                ws.append([s.name, label, date_s, status, content, len(content)])
    else:
        # 没有课次也要给出学生名单，避免导出空文件
        for s in students:
            ws.append([s.name, "（暂无课次）", "", "未生成", "", 0])

    widths = [14, 26, 13, 11, 80, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    # —— Sheet2：完成度汇总（一行 = 一节课）——
    ws2 = wb.create_sheet("完成度汇总")
    ws2.append(["课次", "上课日期", "应写", "已确认", "草稿", "未生成", "完成率"])
    for col in range(1, 8):
        c = ws2.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")

    total_students = len(students)
    for idx, lesson in enumerate(lessons, start=1):
        confirmed = draft = 0
        for s in students:
            r = rev_map.get((lesson.id, s.id))
            if not r:
                continue
            if r.status == "confirmed":
                confirmed += 1
            elif r.status in ("draft", "generating", "pending"):
                draft += 1
        missing = total_students - confirmed - draft
        rate = f"{round(confirmed / total_students * 100)}%" if total_students else "-"
        ws2.append(
            [
                _lesson_label(lesson, idx),
                lesson.lesson_date.strftime("%Y-%m-%d") if lesson.lesson_date else "",
                total_students,
                confirmed,
                draft,
                max(missing, 0),
                rate,
            ]
        )
    for i, w in enumerate([26, 13, 8, 9, 8, 9, 9], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name=f"{klass.name}-课评存档.xlsx",
        as_attachment=True,
    )


@reports_bp.route("/export-pdf/<class_id>", methods=["GET"])
@login_required
def export_pdf(class_id):
    """导出 PDF：按课次分节，逐个学生列出完整课评。"""
    klass, students, lessons, rev_map = _class_scope(class_id)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # 中文字体：优先系统黑体，找不到则回退（英文可读，中文可能显示方块）
    font_name = "Helvetica"
    import os

    for path, name in [
        (r"C:\Windows\Fonts\msyh.ttc", "MSYaHei"),
        (r"C:\Windows\Fonts\simhei.ttf", "SimHei"),
        ("/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc", "WQYZenHei"),
        ("/System/Library/Fonts/PingFang.ttc", "PingFang"),
    ]:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                font_name = name
                break
            except Exception:
                continue

    styles = getSampleStyleSheet()
    st_title = ParagraphStyle(
        "CnTitle", parent=styles["Title"], fontName=font_name, fontSize=18, leading=24
    )
    st_h2 = ParagraphStyle(
        "CnH2",
        parent=styles["Heading2"],
        fontName=font_name,
        fontSize=13,
        leading=18,
        spaceBefore=10,
        textColor="#4F46E5",
    )
    st_name = ParagraphStyle(
        "CnName", parent=styles["Normal"], fontName=font_name, fontSize=11,
        leading=16, spaceBefore=6, textColor="#111827",
    )
    st_body = ParagraphStyle(
        "CnBody", parent=styles["Normal"], fontName=font_name, fontSize=10,
        leading=16, leftIndent=8, textColor="#374151",
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=16 * mm,
        title=f"{klass.name} 课评存档",
    )

    def esc(t):
        return (
            (t or "")
            .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace("\n", "<br/>")
        )

    flow = [
        Paragraph(f"{esc(klass.name)} · 课评存档", st_title),
        Paragraph(
            f"在册学生 {len(students)} 人 ｜ 课次 {len(lessons)} 节", st_body
        ),
        Spacer(1, 10),
    ]

    if not lessons:
        flow.append(Paragraph("该班级暂无课次记录。", st_body))
    for idx, lesson in enumerate(lessons, start=1):
        flow.append(Paragraph(esc(_lesson_label(lesson, idx)), st_h2))
        wrote = False
        for s in students:
            r = rev_map.get((lesson.id, s.id))
            if not r or not (r.content or "").strip():
                continue
            wrote = True
            status = STATUS_CN.get(r.status, r.status)
            flow.append(Paragraph(f"{esc(s.name)}　<font size=8 color='#9ca3af'>[{status}]</font>", st_name))
            flow.append(Paragraph(esc(r.content), st_body))
        if not wrote:
            flow.append(Paragraph("（本次课暂无已生成的课评）", st_body))
        flow.append(Spacer(1, 8))

    doc.build(flow)
    buf.seek(0)
    return send_file(
        buf, mimetype="application/pdf",
        download_name=f"{klass.name}-课评存档.pdf", as_attachment=True,
    )


@reports_bp.route("/reminder")
@login_required
def reminder():
    """未写课评提醒：列出「已建课次但课评没写完」的班级。"""
    klasses = Klass.query.filter_by(
        user_id=current_user.id, deleted_at=None
    ).all()

    pending = []
    for k in klasses:
        stu_n = (
            db.session.query(Enrollment)
            .filter_by(class_id=k.id, user_id=current_user.id)
            .filter(Enrollment.deleted_at.is_(None))
            .count()
        )
        lessons = (
            Lesson.query.filter_by(class_id=k.id, user_id=current_user.id)
            .filter(Lesson.deleted_at.is_(None))
            .all()
        )
        if not lessons or not stu_n:
            continue  # 没课次或没学生的班级不打扰
        need = stu_n * len(lessons)
        done = (
            Review.query.filter_by(class_id=k.id, user_id=current_user.id)
            .filter(Review.deleted_at.is_(None))
            .filter(Review.status.in_(["confirmed", "leave"]))
            .count()
        )
        if done < need:
            pending.append(
                {
                    "class_id": k.id,
                    "class_name": k.name,
                    "need": need,
                    "done": done,
                    "remaining": need - done,
                    "lessons": len(lessons),
                }
            )

    pending.sort(key=lambda x: -x["remaining"])
    return jsonify(
        {
            "total_classes": len(klasses),
            "pending_classes": len(pending),
            "remaining_total": sum(p["remaining"] for p in pending),
            "items": pending[:8],
        }
    )
